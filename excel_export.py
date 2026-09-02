from io import BytesIO

import openpyxl
from openpyxl.styles import Font

from models import NO_TELEGRAM_USERNAME


def _occupation_label(value):
    if value in ("study", "talaba"):
        return "Учится"
    if value in ("work", "ishlayman"):
        return "Работает"
    return value or ""


def _telegram_label(value):
    if value == NO_TELEGRAM_USERNAME:
        return "Нет username"
    return value or ""


def _autosize_columns(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max(12, length + 2)


def build_active_volunteers_workbook(volunteers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Волонтёры"

    headers = ["ФИО", "Телефон", "Telegram", "Пол", "Возраст", "Дата рождения", "Занятость", "Есть авто", "Марка авто", "Номер авто", "Дата регистрации"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for v in volunteers:
        gender_label = "Мужской" if v.gender == "male" else "Женский" if v.gender == "female" else ""
        car_label = "Да" if v.has_car else ("Нет" if v.has_car is False else "Не указано")
        ws.append([
            v.full_name,
            v.phone,
            _telegram_label(v.telegram),
            gender_label,
            v.age or "",
            v.birth_date.strftime("%d.%m.%Y") if v.birth_date else "",
            _occupation_label(v.occupation),
            car_label,
            v.car_brand or "",
            v.car_plate or "",
            v.formatted_date(),
        ])

    _autosize_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_applications_workbook(applications):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["ФИО", "Телефон", "Telegram", "Пол", "Возраст", "Занятость", "Сообщение", "Статус", "Дата подачи"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for a in applications:
        gender_label = "Мужской" if a.gender == "male" else "Женский" if a.gender == "female" else ""
        ws.append([
            a.full_name,
            a.phone,
            _telegram_label(a.telegram),
            gender_label,
            a.age or "",
            _occupation_label(a.occupation),
            a.message or "",
            a.status or "",
            a.created_at.strftime("%d.%m.%Y"),
        ])

    _autosize_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
